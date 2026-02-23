"""
Форматированный экспорт в Excel с красивым оформлением таблиц
"""
import io
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _humanize_header(key: str) -> str:
    """Преобразует snake_case в читаемый заголовок"""
    return key.replace('_', ' ').title()


def create_formatted_excel(
    rows: List[Dict[str, Any]],
    sheet_name: str = "Data",
    currency_columns: Optional[List[str]] = None,
    header_bg: str = "4472C4",
    header_font_color: str = "FFFFFF",
) -> bytes:
    """
    Создаёт отформатированный Excel файл.
    
    Args:
        rows: Список словарей (каждый — строка)
        sheet_name: Имя листа
        currency_columns: Колонки для формата валюты (price, est_project_cost и т.д.)
        header_bg: Цвет фона заголовков (hex без #)
        header_font_color: Цвет шрифта заголовков
    """
    if not rows:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(["No data"])
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel limit

    # Стили
    header_font = Font(bold=True, color=header_font_color, size=11)
    header_fill = PatternFill(start_color=header_bg, end_color=header_bg, fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовки
    headers = list(rows[0].keys())
    header_labels = [_humanize_header(h) for h in headers]
    
    for col_idx, label in enumerate(header_labels, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Данные
    currency_columns = currency_columns or []
    
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, key in enumerate(headers, 1):
            value = row_data.get(key)
            if value is None:
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            if key in currency_columns and isinstance(value, (int, float)):
                cell.number_format = '$#,##0'
                cell.alignment = Alignment(horizontal='right')

    # Ширина колонок
    for col_idx, key in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header_labels[col_idx - 1]))
        for row in rows:
            val = row.get(key)
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 50))
        ws.column_dimensions[col_letter].width = max_len + 2

    # Закрепить первую строку
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
