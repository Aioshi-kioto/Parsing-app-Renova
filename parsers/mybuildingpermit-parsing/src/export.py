"""
Модуль экспорта данных MyBuildingPermit в Excel
Использует pandas и openpyxl для красивого форматирования
"""
from pathlib import Path
from typing import List, Dict, Any
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from config import EXCEL_OUTPUT_DIR, EXCEL_FILENAME


# Стили для Excel
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
OWNER_BUILDER_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def export_to_excel(
    data: List[Dict[str, Any]],
    filename: str = None,
    sheet_name: str = "Permits"
) -> Path:
    """
    Экспортирует данные в Excel с форматированием.
    
    Args:
        data: Список словарей с данными пермитов
        filename: Имя файла (по умолчанию из config)
        sheet_name: Название листа
        
    Returns:
        Path к созданному файлу
    """
    if not data:
        raise ValueError("No data to export")
    
    output_path = EXCEL_OUTPUT_DIR / (filename or EXCEL_FILENAME)
    
    # Определяем порядок колонок
    columns_order = [
        "permit_number",
        "jurisdiction",
        "project_name",
        "description",
        "permit_type",
        "permit_status",
        "address",
        "parcel",
        "applied_date",
        "issued_date",
        "applicant_name",
        "contractor_name",
        "contractor_license",
        "is_owner_builder",
        "permit_url",
    ]
    
    # Заголовки на русском
    column_labels = {
        "permit_number": "Номер пермита",
        "jurisdiction": "Город",
        "project_name": "Название проекта",
        "description": "Описание",
        "permit_type": "Тип",
        "permit_status": "Статус",
        "address": "Адрес",
        "parcel": "Участок",
        "applied_date": "Дата подачи",
        "issued_date": "Дата выдачи",
        "applicant_name": "Заявитель",
        "contractor_name": "Подрядчик",
        "contractor_license": "Лицензия",
        "is_owner_builder": "Owner-Builder",
        "permit_url": "URL",
    }
    
    # Фильтруем колонки, которые есть в данных
    available_columns = [col for col in columns_order if col in data[0]]
    
    # Создаём DataFrame
    df = pd.DataFrame(data)
    df = df.reindex(columns=available_columns)
    
    # Переименовываем колонки
    df.columns = [column_labels.get(col, col) for col in df.columns]
    
    # Сохраняем в Excel
    df.to_excel(output_path, index=False, sheet_name=sheet_name, engine="openpyxl")
    
    # Применяем форматирование
    _apply_formatting(output_path, sheet_name, len(df))
    
    return output_path


def _apply_formatting(filepath: Path, sheet_name: str, row_count: int) -> None:
    """Применяет стили к Excel файлу"""
    wb = load_workbook(filepath)
    ws = wb[sheet_name]
    
    # Форматируем заголовки
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    
    # Авто-ширина колонок
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        for cell in column_cells:
            try:
                cell_value = str(cell.value) if cell.value else ""
                max_length = max(max_length, len(cell_value))
            except:
                pass
        
        # Ограничиваем ширину
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Подсветка Owner-Builder = TRUE
    owner_builder_col = None
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value == "Owner-Builder":
            owner_builder_col = col_idx
            break
    
    if owner_builder_col:
        for row_idx in range(2, row_count + 2):
            cell = ws.cell(row=row_idx, column=owner_builder_col)
            if cell.value is True or str(cell.value).upper() == "TRUE":
                cell.fill = OWNER_BUILDER_FILL
                cell.font = Font(bold=True, color="006100")
    
    # Форматируем URL как гиперссылки
    url_col = None
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value == "URL":
            url_col = col_idx
            break
    
    if url_col:
        for row_idx in range(2, row_count + 2):
            cell = ws.cell(row=row_idx, column=url_col)
            if cell.value and str(cell.value).startswith("http"):
                cell.hyperlink = str(cell.value)
                cell.font = Font(color="0563C1", underline="single")
    
    # Сохраняем
    wb.save(filepath)
    wb.close()


def create_summary_report(data: List[Dict[str, Any]], output_path: Path = None) -> Path:
    """
    Создаёт сводный отчёт по городам
    """
    if not output_path:
        output_path = EXCEL_OUTPUT_DIR / f"permits_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    df = pd.DataFrame(data)
    
    # Сводка по городам
    summary = df.groupby("jurisdiction").agg({
        "permit_number": "count",
        "is_owner_builder": lambda x: sum(x == True) if "is_owner_builder" in df.columns else 0
    }).rename(columns={
        "permit_number": "Всего пермитов",
        "is_owner_builder": "Owner-Builder"
    })
    
    summary.to_excel(output_path, sheet_name="Сводка")
    
    return output_path


if __name__ == "__main__":
    # Тест экспорта
    test_data = [
        {
            "permit_number": "BFL26-0008",
            "jurisdiction": "Auburn",
            "project_name": "Test Project",
            "address": "123 Main St",
            "permit_type": "BUILDING",
            "permit_status": "IN REVIEW",
            "applicant_name": "John Doe",
            "contractor_name": "",
            "is_owner_builder": True,
            "permit_url": "https://permitsearch.mybuildingpermit.com/PermitDetails/BFL26-0008/Auburn",
        }
    ]
    
    output = export_to_excel(test_data, "test_export.xlsx")
    print(f"Exported to: {output}")
