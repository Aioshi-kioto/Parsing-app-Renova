"""
Convert main CSV files to Excel for review.
- Keeps only fields defined in docs/master_spec.md (section 7)
- Applies formatting: column widths, header style, freeze panes
Output: output/seattle_permits_review.xlsx
"""
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Fields to keep (API lowercase -> Excel column name). All other fields are removed.
FIELDS_TO_KEEP = {
    "permitnum": "PermitNum",
    "permitclass": "PermitClass",
    "permitclassmapped": "PermitClassMapped",
    "permittypemapped": "PermitTypeMapped",
    "permittypedesc": "PermitTypeDesc",
    "description": "Description",
    "estprojectcost": "EstProjectCost",
    "applieddate": "AppliedDate",
    "originaladdress1": "OriginalAddress1",
    "originalcity": "OriginalCity",
    "originalstate": "OriginalState",
    "originalzip": "OriginalZip",
    "contractorcompanyname": "ContractorCompanyName",
    "link": "Link",
    "housingcategory": "HousingCategory",
}


def filter_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Leave only FIELDS_TO_KEEP, rename to display names."""
    df_lower = df.copy()
    df_lower.columns = df_lower.columns.str.lower()

    keep = {k: v for k, v in FIELDS_TO_KEEP.items() if k in df_lower.columns}
    out = df_lower[[k for k in keep]].copy()
    out.columns = [keep[c] for c in out.columns]
    return out


def style_worksheet(ws):
    """Header: bold, background; column widths; freeze first row."""
    thin = Side(style="thin", color="CCCCCC")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Column widths (approximate)
    widths = {
        "A": 14,   # PermitNum
        "B": 22,   # PermitClass
        "C": 18,   # PermitClassMapped
        "D": 14,   # PermitTypeMapped
        "E": 28,   # PermitTypeDesc
        "F": 50,   # Description
        "G": 14,   # EstProjectCost
        "H": 12,   # AppliedDate
        "I": 28,   # OriginalAddress1
        "J": 12,   # OriginalCity
        "K": 10,   # OriginalState
        "L": 10,   # OriginalZip
        "M": 22,   # ContractorCompanyName
        "N": 55,   # Link
        "O": 20,   # HousingCategory
    }
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    # Number format for cost and date
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=7).number_format = "#,##0"      # EstProjectCost
        ws.cell(row=row, column=8).number_format = "YYYY-MM-DD"   # AppliedDate

    ws.freeze_panes = "A2"
    if ws.max_row > 1 and ws.max_column >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def main():
    output_dir = Path("output")
    # Use timestamp so we don't overwrite open file
    from datetime import datetime
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    excel_file = output_dir / f"seattle_permits_review_{ts}.xlsx"

    files = [
        ("owner_builder_candidates_2025_20260129_015346.csv", "Owner-Builder Candidates"),
        ("all_single_family_2025_20260129_015346.csv", "All Single Family 2025"),
        ("master_owner_builders.csv", "Master Daily Update"),
    ]

    with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
        for filename, sheet_name in files:
            path = output_dir / filename
            if path.exists():
                df = pd.read_csv(path)
                df_out = filter_columns(df)
                sheet_name_safe = sheet_name[:31].replace("/", "-").replace("\\", "-")
                df_out.to_excel(writer, sheet_name=sheet_name_safe, index=False)
                print(f"  {filename} -> sheet '{sheet_name_safe}' ({len(df_out)} rows, {len(df_out.columns)} cols)")

    # Apply formatting to all sheets
    wb = load_workbook(excel_file)
    for sheet_name in wb.sheetnames:
        style_worksheet(wb[sheet_name])
    wb.save(excel_file)

    print(f"\nSaved: {excel_file}")
    print("Columns: " + ", ".join(FIELDS_TO_KEEP.values()))


if __name__ == "__main__":
    main()
